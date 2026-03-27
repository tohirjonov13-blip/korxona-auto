"""
app.py — 1-Korxona Автозаполнение
Единый интерфейс: загрузка → проверки → ручные поля → скачать отчёт
Запуск: streamlit run app.py
"""

import streamlit as st
import pandas as pd
import tempfile, os, io, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _make_excel(proc, ndfl, tmp_dir):
    """Формирует Excel-отчёт без шаблона"""
    try:
        from processor import MAPPING
    except ImportError:
        MAPPING = {}

    thin = Side(style="thin", color="CCCCCC")
    BD   = Border(left=thin, right=thin, top=thin, bottom=thin)
    WB   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DB   = Font(name="Arial", bold=True, color="1F3864", size=9)
    DR   = Font(name="Arial", color="1F3864", size=9)
    SR   = Font(name="Arial", color="555555", size=8)
    DARK = PatternFill("solid", fgColor="1F3864")
    MED  = PatternFill("solid", fgColor="2E75B6")
    LBLU = PatternFill("solid", fgColor="D6E4F0")
    GRAY = PatternFill("solid", fgColor="F5F5F5")
    C    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    R    = Alignment(horizontal="right",  vertical="center")

    def H(ws,r,c,v,f=MED):
        cl=ws.cell(r,c,v); cl.fill=f; cl.border=BD; cl.font=WB; cl.alignment=C
    def Ce(ws,r,c,v,f=None,fo=DR,al=L):
        cl=ws.cell(r,c,v)
        if f: cl.fill=f
        cl.border=BD; cl.font=fo; cl.alignment=al
    def Nu(ws,r,c,v,f=None):
        cl=ws.cell(r,c,float(v) if v else 0)
        if f: cl.fill=f
        cl.border=BD; cl.font=DR; cl.alignment=R; cl.number_format="#,##0"

    wb = openpyxl.Workbook()

    # ── Главный лист ──────────────────────────────────────────────────
    ws = wb.active; ws.title="1-korxona"
    ws.column_dimensions["A"].width=8
    ws.column_dimensions["B"].width=46
    ws.column_dimensions["C"].width=20
    ws.column_dimensions["D"].width=18
    ws.column_dimensions["E"].width=18

    inn = (ndfl.inn if ndfl else "—")
    ws.merge_cells("A1:E1")
    cl=ws["A1"]; cl.value=f"1-KORXONA | ИНН {inn} | 2025 год"
    cl.fill=DARK; cl.font=WB; cl.alignment=C; ws.row_dimensions[1].height=28

    H(ws,2,1,"Код"); H(ws,2,2,"Показатель")
    H(ws,2,3,"Значение"); H(ws,2,4,"Нач. года"); H(ws,2,5,"Кон. года")
    ws.row_dimensions[2].height=32

    CHAPTERS = {
        "Глава 1 — Доходы":            list(range(100,110)),
        "Глава 2 — Затраты":           list(range(110,128)),
        "Глава 3 — Запасы":            list(range(140,146)),
        "Глава 4 — ИКТ":               list(range(150,153)),
        "Глава 5 — Основные средства": list(range(160,172)),
        "Глава 6 — Инвестиции":        list(range(180,187)),
        "Глава 8 — Энергоресурсы":     list(range(301,310)),
        "Глава 9 — Кадры":             list(range(401,417)),
        "Глава 10 — Выплаты":          list(range(417,425)),
    }

    row=3
    for chapter, codes in CHAPTERS.items():
        ws.merge_cells(f"A{row}:E{row}")
        cl=ws.cell(row,1,chapter); cl.fill=MED; cl.font=WB
        cl.alignment=L; cl.border=BD; ws.row_dimensions[row].height=20
        row+=1
        for code in codes:
            res=proc.results.get(code)
            if res is None: continue
            cfg=MAPPING.get(code,{})
            fl=LBLU if row%2==0 else GRAY
            Ce(ws,row,1,code,fl,SR,C)
            Ce(ws,row,2,cfg.get("desc",f"Код {code}"),fl,DB)
            yv=res.get("year"); bv=res.get("begin"); ev=res.get("end")
            if yv is not None: Nu(ws,row,3,yv,fl)
            else: Ce(ws,row,3,"—",fl,SR,C)
            if bv is not None: Nu(ws,row,4,bv,fl)
            else: Ce(ws,row,4,"—",fl,SR,C)
            if ev is not None: Nu(ws,row,5,ev,fl)
            else: Ce(ws,row,5,"—",fl,SR,C)
            ws.row_dimensions[row].height=15
            row+=1

    # ── Лист НДФЛ ─────────────────────────────────────────────────────
    if ndfl:
        ws2=wb.create_sheet("НДФЛ")
        ws2.column_dimensions["A"].width=42; ws2.column_dimensions["B"].width=22
        ws2.merge_cells("A1:B1")
        cl=ws2["A1"]; cl.value="НДФЛ-РАСЧЁТ"; cl.fill=DARK; cl.font=WB; cl.alignment=C
        rows=[
            ("Общие доходы (010)",      ndfl.calc.total_income),
            ("Доходы ОТ (011)",          ndfl.calc.labor_income),
            ("  ЗП в периоде (0110)",    ndfl.calc.salary_period),
            ("Доходы не ОТ (012)",       ndfl.calc.non_labor_income),
            ("Освобождённые (030)",      ndfl.calc.exempt_income),
            ("НДФЛ начисл. (060)",       ndfl.calc.ndfl_accrued),
            ("Итого НДФЛ+СН (070)",      ndfl.calc.total_tax),
            ("Сотрудников (Прил.4)",     len(ndfl.employees)),
            ("Призовых (Прил.5)",        len(ndfl.prize_employees)),
        ]
        for i,(lbl,val) in enumerate(rows):
            r=i+2; fl=LBLU if i%2==0 else GRAY
            Ce(ws2,r,1,lbl,fl,DB if not lbl.startswith(" ") else DR)
            Nu(ws2,r,2,val,fl)
            ws2.row_dimensions[r].height=16

        # Список сотрудников
        ws3=wb.create_sheet("Сотрудники")
        for col,w in zip("ABCDEFGHIJ",[4,36,14,18,10,10,12,18,16,14]):
            ws3.column_dimensions[col].width=w
        ws3.merge_cells("A1:J1")
        cl=ws3["A1"]; cl.value="ПРИЛОЖЕНИЕ 4 — Сотрудники"; cl.fill=DARK; cl.font=WB; cl.alignment=C
        hdrs=["№","Ф.И.О.","Должность","ПИНФЛ","Резидент","Статус","Контракт","Доход (сум)","НДФЛ (сум)","Ставка"]
        for i,h in enumerate(hdrs,1): H(ws3,2,i,h)
        CMAP={1:"Основной",2:"Совместитель",3:"ГПХ","1":"Основной","2":"Совместитель","3":"ГПХ"}
        SMAP={1:"Работает",2:"Уволен","1":"Работает","2":"Уволен"}
        RMAP={1:"Резидент",2:"Нерезидент","1":"Резидент","2":"Нерезидент"}
        RED=PatternFill("solid",fgColor="FDDEDE")
        YEL=PatternFill("solid",fgColor="FFF9C4")
        for i,e in enumerate(ndfl.employees):
            r=i+3
            fl=(RED if e.is_fired else YEL if e.is_nonresident else
                PatternFill("solid",fgColor="D6E4F0") if e.is_gph else
                GRAY if i%2==0 else None)
            Ce(ws3,r,1,e.num,fl,SR,C); Ce(ws3,r,2,e.name,fl,DB)
            Ce(ws3,r,3,e.position,fl,SR); Ce(ws3,r,4,e.pinfl,fl,SR,C)
            Ce(ws3,r,5,RMAP.get(e.resident,"?"),fl,SR,C)
            Ce(ws3,r,6,SMAP.get(e.status,"?"),fl,SR,C)
            Ce(ws3,r,7,CMAP.get(e.contract,"?"),fl,SR,C)
            Nu(ws3,r,8,e.total_income,fl); Nu(ws3,r,9,e.ndfl_total,fl)
            Ce(ws3,r,10,e.work_rate,fl,SR,C)
            ws3.row_dimensions[r].height=14

    path=os.path.join(tmp_dir,"1korxona.xlsx")
    wb.save(path); return path



st.set_page_config(
    page_title="1-Korxona · Автозаполнение",
    page_icon="📊",
    layout="centered",
)

st.markdown("""
<style>
/* ── Шапка — тёмная полоса, работает в обоих режимах ── */
.app-header {
    background: #1a3c5e;
    padding: 1.2rem 1.5rem;
    border-radius: 10px;
    margin-bottom: 1.5rem;
    display: flex;
    align-items: center;
    gap: 12px;
}
.app-header h1 { margin:0; font-size:1.25rem; font-weight:600; color:#fff; }
.app-header p  { margin:.2rem 0 0; font-size:.82rem; color:#93c5e8; }

/* ── Заголовки шагов ── */
.step-hdr {
    font-size: .82rem;
    font-weight: 600;
    color: var(--text-color);
    text-transform: uppercase;
    letter-spacing: .06em;
    margin: 1.4rem 0 .7rem;
    display: flex;
    align-items: center;
    gap: 8px;
    opacity: .7;
}
.step-hdr::after {
    content: '';
    flex: 1;
    height: 0.5px;
    background: currentColor;
    opacity: .2;
}

/* ── Карточки проверок — семантические CSS-переменные Streamlit ── */
.chk-ok, .chk-warn, .chk-err {
    display: flex;
    gap: 10px;
    align-items: flex-start;
    border-radius: 8px;
    padding: .7rem .9rem;
    margin: .4rem 0;
    font-size: .88rem;
    line-height: 1.5;
    border: 0.5px solid;
}
.chk-ok   {
    background: rgba(34,197,94,.12);
    border-color: rgba(34,197,94,.35);
    color: var(--text-color);
}
.chk-warn {
    background: rgba(245,158,11,.12);
    border-color: rgba(245,158,11,.35);
    color: var(--text-color);
}
.chk-err  {
    background: rgba(239,68,68,.12);
    border-color: rgba(239,68,68,.35);
    color: var(--text-color);
}

/* Иконки — фиксированные цвета, читаются в любом режиме */
.chk-icon {
    width: 20px; height: 20px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: 11px; font-weight: 700; flex-shrink: 0; margin-top: 1px;
    color: #fff;
}
.chk-ok   .chk-icon { background: #22c55e; }
.chk-warn .chk-icon { background: #f59e0b; }
.chk-err  .chk-icon { background: #ef4444; }

.chk-title { font-weight: 600; font-size: .88rem; }
.chk-desc  { font-size: .82rem; opacity: .75; margin-top: 2px; }
.chk-tip   {
    font-size: .78rem;
    color: #3b82f6;
    margin-top: 4px;
}
</style>
""", unsafe_allow_html=True)

# ── Шапка ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-header">
  <div style="font-size:1.6rem">📊</div>
  <div>
    <h1>1-Korxona · Автозаполнение</h1>
    <p>Статистический отчёт РУз · Загрузите файлы из 1С</p>
  </div>
</div>
""", unsafe_allow_html=True)

# ─── session state ────────────────────────────────────────────────────────────
for k, v in dict(parsed=False, checked=False, skip_warn=False,
                 proc=None, ndfl=None, pers=None, summary=None).items():
    if k not in st.session_state:
        st.session_state[k] = v
S = st.session_state

# ══════════════════════════════════════════════════════════════════════
# ШАГ 1 — ЗАГРУЗКА
# ══════════════════════════════════════════════════════════════════════
st.markdown('<div class="step-hdr"><span>①</span> Загрузите файлы из 1С</div>',
            unsafe_allow_html=True)

c1, c2 = st.columns(2)
with c1:
    st.caption("**Финансовая отчётность**")
    f1_f   = st.file_uploader("Форма 1 (Баланс)",       type=["xlsx","xls","xltx"], key="u1")
    f2_f   = st.file_uploader("Форма 2 (ОПУ)",          type=["xlsx","xls","xltx"], key="u2")
    osv_f  = st.file_uploader("ОСВ из 1С",              type=["xlsx","xls"],        key="u8",
                               help="Оборотно-сальдовая ведомость за год. Нужна для кодов: амортизация, материалы, ФОТ, движение ОС")
    ndfl_f = st.file_uploader("Расчёт НДФЛ (годовой)",  type=["xlsx","xls"],        key="u3")
with c2:
    st.caption("**HR-данные**")
    gph_f  = st.file_uploader("Список ГПХ",             type=["xlsx","xls"],        key="u4")
    hire_f = st.file_uploader("Список приёма",           type=["xlsx","xls"],        key="u5")
    fire_f = st.file_uploader("Список увольнений",       type=["xlsx","xls"],        key="u6")

tmpl_f = st.file_uploader(
    "Шаблон 1-korxona от Госкомстата (необязательно — если загружен, заполним его)",
    type=["xlsx","xls"], key="u7"
)

# Показываем что загружено
loaded_names = []
if f1_f:   loaded_names.append("✅ Форма 1")
if f2_f:   loaded_names.append("✅ Форма 2")
if osv_f:  loaded_names.append("✅ ОСВ")
if ndfl_f: loaded_names.append("✅ НДФЛ")
if gph_f:  loaded_names.append("✅ ГПХ")
if hire_f: loaded_names.append("✅ Приём")
if fire_f: loaded_names.append("✅ Увольнения")
if loaded_names:
    st.caption("Загружено: " + "  ".join(loaded_names))

if not any([f1_f, f2_f, ndfl_f]):
    st.info("Загрузите хотя бы Форму 1, Форму 2 или Расчёт НДФЛ.")
    st.stop()

# Сохраняем файлы
tmp = tempfile.mkdtemp()
def sv(f, name):
    if not f: return None
    p = os.path.join(tmp, name)
    open(p, "wb").write(f.getbuffer())
    return p

# ОСВ сохраняем с оригинальным расширением (важно для .xls конвертации)
def sv_osv(f):
    if not f: return None
    ext = f.name.rsplit(".", 1)[-1].lower()
    p = os.path.join(tmp, f"osv.{ext}")
    open(p, "wb").write(f.getbuffer())
    return p

P = {
    "f1":   sv(f1_f,   "f1.xlsx"),
    "f2":   sv(f2_f,   "f2.xlsx"),
    "osv":  sv_osv(osv_f),
    "ndfl": sv(ndfl_f, "ndfl.xlsx"),
    "gph":  sv(gph_f,  "gph.xlsx"),
    "hire": sv(hire_f, "hire.xlsx"),
    "fire": sv(fire_f, "fire.xlsx"),
    "tmpl": sv(tmpl_f, "tmpl.xlsx"),
}

if st.button("▶ Обработать файлы", type="primary", use_container_width=True):
    S.parsed = S.checked = S.skip_warn = False
    S.proc = S.ndfl = S.pers = S.summary = None

# ══════════════════════════════════════════════════════════════════════
# ПАРСИНГ
# ══════════════════════════════════════════════════════════════════════
if not S.parsed and any([P["f1"], P["f2"], P["ndfl"]]):
    with st.spinner("Читаем файлы…"):
        try:
            from processor import KorxonaProcessor
            proc = KorxonaProcessor(f1_path=P["f1"], f2_path=P["f2"], osv_path=P["osv"])
            if P["f1"]:  proc.parse_f1()
            if P["f2"]:  proc.parse_f2()
            if P["osv"]: proc.parse_osv()
            proc.compute()

            if P["ndfl"]:
                from ndfl_processor import (parse_ndfl_report, parse_gph_list,
                    parse_hire_list, parse_fire_list, extract_korxona_personnel)
                from ndfl_checks import run_all_checks
                ndfl = parse_ndfl_report(P["ndfl"])
                gph  = parse_gph_list(P["gph"])  if P["gph"]  else []
                dfh  = parse_hire_list(P["hire"]) if P["hire"] else pd.DataFrame()
                dff  = parse_fire_list(P["fire"]) if P["fire"] else pd.DataFrame()
                pers = extract_korxona_personnel(ndfl, dff if not dff.empty else None)
                for code, d in pers.items():
                    proc.results[code] = {"year": d["value"]}
                S.ndfl = ndfl
                S.pers = pers
                S.summary = run_all_checks(ndfl, gph, dfh, dff, pers)

            S.proc   = proc
            S.parsed = True
        except Exception as e:
            st.error(f"Ошибка: {e}")
            import traceback
            with st.expander("Детали"):
                st.code(traceback.format_exc())
            st.stop()

if not S.parsed:
    st.stop()

# ══════════════════════════════════════════════════════════════════════
# ШАГ 2 — ПРОВЕРКИ
# ══════════════════════════════════════════════════════════════════════
st.divider()
st.markdown('<div class="step-hdr"><span>②</span> Результаты проверок</div>',
            unsafe_allow_html=True)

if S.summary is None:
    st.info("НДФЛ не загружен — проверки пропущены.")
    S.checked = True
else:
    from ndfl_checks import Severity
    crit = S.summary.critical
    warns = S.summary.warnings

    col_a, col_b, col_c = st.columns(3)
    col_a.metric("❌ Критичных",      len(crit))
    col_b.metric("⚠️ Предупреждений", len(warns))
    col_c.metric("✅ Ок",             len(S.summary.infos))

    problems = [r for r in S.summary.results if r.severity != Severity.INFO]
    if not problems:
        st.markdown('<div class="chk-ok">✅ Все проверки пройдены</div>',
                    unsafe_allow_html=True)
        S.checked = True
    else:
        for r in problems:
            if r.severity == Severity.CRITICAL:
                css, icon_char, icon_bg = "chk-err", "✕", "#ef4444"
            else:
                css, icon_char, icon_bg = "chk-warn", "!", "#f59e0b"
            rec = (f'<div class="chk-tip">💡 {r.recommendation}</div>'
                   if r.recommendation else "")
            st.markdown(f"""
            <div class="{css}">
              <div class="chk-icon" style="background:{icon_bg}">{icon_char}</div>
              <div>
                <div class="chk-title">{r.title}</div>
                <div class="chk-desc">{r.description}</div>
                {rec}
              </div>
            </div>""", unsafe_allow_html=True)
            if r.affected:
                with st.expander(f"Подробнее ({len(r.affected)})"):
                    for x in r.affected: st.markdown(f"• {x}")

        blocking = [r for r in crit if not r.can_skip]
        if blocking:
            st.error("Исправьте критичные ошибки перед продолжением.")
            st.stop()
        else:
            if not S.skip_warn:
                if st.button("Пропустить предупреждения и продолжить →",
                             use_container_width=True):
                    S.skip_warn = True
                    S.checked   = True
                    st.rerun()
                st.stop()
            else:
                S.checked = True

if not S.checked:
    st.stop()

# ══════════════════════════════════════════════════════════════════════
# ШАГ 3 — РУЧНЫЕ ПОЛЯ
# ══════════════════════════════════════════════════════════════════════
st.divider()
st.markdown('<div class="step-hdr"><span>③</span> Дополните данные вручную</div>',
            unsafe_allow_html=True)

proc = S.proc
ndfl = S.ndfl
pers = S.pers or {}

def get_val(code):
    """Текущее значение из processor"""
    if not proc: return 0.0
    r = proc.results.get(code, {})
    return float(r.get("year") or 0)

def set_val(code, val):
    if proc: proc.results[code] = {"year": val}

# Глава 9 — Кадры
with st.expander("📋 Глава 9 — Кадры (проверьте автозаполненные данные)", expanded=True):
    st.caption("Значения заполнены автоматически из НДФЛ. Скорректируйте если нужно.")
    g9 = st.columns(3)
    G9 = [(401,"Числ. для ЗП",1),(403,"ФОТ (сум)",1000),
          (409,"Ср.год. с тр.кн.",1),(411,"Совместители",1),
          (412,"ГПХ-работники",1),(413,"Итого числ.",1),
          (404,"ФОТ с тр.кн.",1000),(405,"Числ. на конец года",1),
          (416,"Всего расходов",1000)]
    for i, (code, lbl, step) in enumerate(G9):
        with g9[i % 3]:
            v = st.number_input(f"Код {code} — {lbl}",
                value=get_val(code), min_value=0.0,
                step=float(step), key=f"g9_{code}", format="%.0f")
            set_val(code, v)
    # Контроль 413
    v409,v411,v412,v413 = [get_val(c) for c in (409,411,412,413)]
    exp = v409+v411+v412
    if abs(v413-exp) > 0 and (v409+v411+v412) > 0:
        st.warning(f"413={v413:.0f} ≠ 409+411+412={exp:.0f}")
        if st.button("Исправить 413 = "+str(int(exp))):
            st.session_state["g9_413"] = exp
            set_val(413, exp)
            st.rerun()

# Глава 10 — Выплаты
with st.expander("💰 Глава 10 — Выплаты физлицам (тыс. сум)", expanded=True):
    g10 = st.columns(2)
    G10 = [(417,"Проценты (ст.375)"),(418,"Дивиденды (ст.375)"),
           (419,"Матер. выгода (ст.376)"),(420,"Матер. помощь (ст.378)"),
           (421,"Авторское вознагр. (ст.393)"),(422,"Выходное пособие (ст.377)"),
           (423,"Компенсации (ст.369)"),(424,"Обучение сотрудников")]
    for i, (code, lbl) in enumerate(G10):
        with g10[i % 2]:
            v = st.number_input(f"Код {code} — {lbl}",
                value=get_val(code), min_value=0.0,
                step=1000.0, key=f"g10_{code}", format="%.0f")
            set_val(code, v)

# Глава 4 — ИКТ
with st.expander("💻 Глава 4 — Затраты на ИКТ (тыс. сум)"):
    g4 = st.columns(3)
    G4 = [(150,"ИКТ — всего"),(151,"Программное обеспечение"),(152,"Хостинг")]
    for i, (code, lbl) in enumerate(G4):
        with g4[i]:
            v = st.number_input(f"Код {code} — {lbl}",
                value=get_val(code), min_value=0.0,
                step=100.0, key=f"g4_{code}", format="%.0f")
            set_val(code, v)

# Глава 8 — Энергоресурсы
with st.expander("⚡ Глава 8 — Энергоресурсы (если применимо)"):
    g8 = st.columns(2)
    G8 = [(301,"Природный газ (тыс.куб.м)"),(302,"Электроэнергия (тыс.кВт/ч)"),
          (305,"Дизельное топливо (тонн)"),(306,"Бензин (тонн)")]
    for i, (code, lbl) in enumerate(G8):
        with g8[i % 2]:
            v = st.number_input(f"Код {code}",
                value=get_val(code), min_value=0.0,
                step=0.1, key=f"g8_{code}", format="%.2f")
            set_val(code, v)

# ══════════════════════════════════════════════════════════════════════
# ШАГ 4 — ГЕНЕРАЦИЯ
# ══════════════════════════════════════════════════════════════════════
st.divider()
st.markdown('<div class="step-hdr"><span>④</span> Скачать готовый отчёт</div>',
            unsafe_allow_html=True)

if ndfl:
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Сотрудников", len(ndfl.employees))
    m2.metric("Призы",       len(ndfl.prize_employees))
    m3.metric("ФОТ (тыс)",   f"{ndfl.calc.labor_income/1000:,.0f}")
    m4.metric("НДФЛ (тыс)",  f"{ndfl.calc.total_tax/1000:,.0f}")

if st.button("⬇ Сформировать 1-korxona.xlsx", type="primary",
             use_container_width=True):
    with st.spinner("Генерируем…"):
        try:
            out = os.path.join(tmp, "out.xlsx")
            if P["tmpl"]:
                proc.fill_template(P["tmpl"], out)
            else:
                out = _make_excel(proc, ndfl, tmp)
            st.download_button(
                "⬇ Скачать 1-korxona.xlsx",
                data=open(out,"rb").read(),
                file_name="1korxona.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary", use_container_width=True,
            )
            st.success("✅ Готово!")
        except Exception as e:
            st.error(f"Ошибка: {e}")
            import traceback
            with st.expander("Детали"): st.code(traceback.format_exc())
