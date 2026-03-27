"""
processor.py — Основной модуль обработки финансовых данных для отчёта 1-korxona
Поддерживает: Форма 1 (Баланс), Форма 2 (ОПУ), ОСВ из 1С (Excel .xlsx)
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from pathlib import Path
import logging
import re

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────
# МАППИНГ: код строки → источник данных
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────
# МАППИНГ по шпаргалке TAVAT consulting (проверено на реальных файлах)
#
# Источники:
#   f1       — Форма 1 (Баланс), строка берётся из col "both" (нач/кон)
#   f2       — Форма 2 (ОПУ), столбец 5 = доходы, столбец 6 = расходы
#   f2_sum   — Сумма нескольких строк Ф2
#   f2_exp   — Строка Ф2, столбец расходов (col 6 в xltx)
#   osv_dt   — ОСВ, оборот Дебет счёта (÷1000 при osv_div1000=True)
#   osv_kt   — ОСВ, оборот Кредит счёта (÷1000 при osv_div1000=True)
#   osv_beg  — ОСВ, сальдо начало (begin_debit)
#   osv_end  — ОСВ, сальдо конец (end_debit)
#   ndfl     — НДФЛ-расчёт (строка, колонка, ÷1000)
#   calc     — Расчётная формула из других кодов
#   manual   — Ручной ввод бухгалтером
# ─────────────────────────────────────────────────────────────────────

MAPPING = {
    # ══ Глава 1: Доходы ═══════════════════════════════════════════════
    # 101 = Ф2 стр.010 столбец 5 (чистая выручка, доходы)
    101: {"source": "f2",     "row_code": "010", "col": "income",
          "desc": "Оборот организации (без НДС)"},

    # 102 = Ф2 стр.020 столбец 5 (продукция собственного потребления)
    102: {"source": "f2",     "row_code": "020", "col": "income",
          "desc": "Продукция собственного производства для внутр. нужд"},

    # 103 = Ф2 стр.090 + стр.110 столбец 5 (прочие операционные доходы)
    103: {"source": "f2_sum", "row_codes": ["090", "110"], "col": "income",
          "desc": "Прочие доходы"},

    # 104 = ОСВ счёт 9300/93 кредит оборот (доходы от аренды)
    104: {"source": "osv_kt", "accounts": ["9300", "930", "93"],
          "osv_div1000": True,
          "desc": "Доходы от оперативной аренды"},

    # 105 = Ф2 стр.120 столбец 5 (дивиденды полученные)
    105: {"source": "f2",     "row_code": "120", "col": "income",
          "desc": "Доходы в виде дивидендов"},

    # 106 = Ф2 стр.130 столбец 5 (проценты полученные)
    106: {"source": "f2",     "row_code": "130", "col": "income",
          "desc": "Целевые поступления"},

    # 109 = 101 + 102 + 103 (контрольное соотношение)
    109: {"source": "calc", "formula": "101+102+103",
          "desc": "Общие доходы (101+102+103)"},

    # ══ Глава 2: Затраты ══════════════════════════════════════════════
    # 110 = 111 + 127 (затраты всего = себест.+расходы + финансовые расходы)
    110: {"source": "calc", "formula": "111+127",
          "desc": "Затраты – всего"},

    # 111 = Ф2 стр.020 + стр.040 столбец 6 (себестоимость + расходы периода)
    111: {"source": "f2_sum_exp", "row_codes": ["020", "040"], "col": "expense",
          "desc": "Себестоимость и расходы периода"},

    # 112 = ОСВ счёт 9120 Дебет оборот (товары для перепродажи)
    #       Если субсчёт 9120 отсутствует — берём 9100
    112: {"source": "osv_dt", "accounts": ["9120", "9100"],
          "osv_div1000": True,
          "desc": "Стоимость товаров для перепродажи (сч.9120)"},

    # 113 = ОСВ счёт 1000 Дебет оборот (материальные затраты)
    113: {"source": "osv_dt", "accounts": ["1000", "1010", "1020", "1030"],
          "osv_div1000": True,
          "desc": "Материальные затраты"},

    # 117 = НДФЛ стр.011 колонка 3 ÷ 1000 (ФОТ начисленный)
    117: {"source": "ndfl",   "ndfl_row": "011", "ndfl_col": 3, "ndfl_div1000": True,
          "desc": "Затраты на оплату труда"},

    # 118 = НДФЛ стр.060 колонка 4 ÷ 1000  ИЛИ  ОСВ сч.6500 Дт оборот ÷ 1000
    118: {"source": "osv_dt", "accounts": ["6500", "6510", "6520"],
          "osv_div1000": True,
          "desc": "Отчисления на социальное страхование (ЕСН)"},

    # 119 = ОСВ счёт 0200 Кредит оборот ÷ 1000 (амортизация ОС)
    119: {"source": "osv_kt", "accounts": ["0200", "0210"],
          "osv_div1000": True,
          "desc": "Амортизация ОС и НМА"},

    # 127 = Ф2 стр.170 столбец 6 (расходы от финансовой деятельности)
    127: {"source": "f2",     "row_code": "170", "col": "expense",
          "desc": "Расходы по финансовой деятельности"},

    # ══ Глава 3: Запасы ═══════════════════════════════════════════════
    # 140 = Ф1 стр.150 нач/кон (производственные запасы)
    140: {"source": "f1", "row_code": "150", "col": "both",
          "desc": "Производственные запасы"},

    # 141 = Ф1 стр.160 нач/кон (незавершённое производство)
    141: {"source": "f1", "row_code": "160", "col": "both",
          "desc": "Незавершённое производство"},

    # 142 = Ф1 стр.170 нач/кон (готовая продукция)
    142: {"source": "f1", "row_code": "170", "col": "both",
          "desc": "Готовая продукция (по рыночным ценам)"},

    # 143 = Ф1 стр.180 нач/кон (товары по себестоимости)
    143: {"source": "f1", "row_code": "180", "col": "both",
          "desc": "Товары (в закупочных ценах)"},

    # ══ Глава 4: ИКТ ══════════════════════════════════════════════════
    150: {"source": "manual", "desc": "Затраты на ИКТ – всего"},
    151: {"source": "manual", "desc": "  из них на программное обеспечение"},
    152: {"source": "manual", "desc": "  за услуги хостинга"},

    # ══ Глава 5: Основные средства ════════════════════════════════════
    # 160 = Ф1 стр.010 нач/кон (ОС первоначальная стоимость)
    160: {"source": "f1", "row_code": "010", "col": "both",
          "desc": "ОС по первоначальной стоимости"},

    # 161 = Ф1 стр.011 нач/кон (износ ОС)
    161: {"source": "f1", "row_code": "011", "col": "both",
          "desc": "Сумма износа ОС"},

    # 162 = Ф1 стр.090 + стр.100 нач/кон (незавершённое строительство)
    162: {"source": "f1_sum", "row_codes": ["090", "100"], "col": "both",
          "desc": "Незавершённое строительство"},

    # 163 = Ф1 стр.020 нач/кон (НМА первоначальная стоимость)
    163: {"source": "f1", "row_code": "020", "col": "both",
          "desc": "НМА по первоначальной стоимости"},

    # 164 = Ф1 стр.021 нач/кон (амортизация НМА)
    164: {"source": "f1", "row_code": "021", "col": "both",
          "desc": "Сумма амортизации НМА"},

    # 165 = ОСВ счёт 0100 Кредит оборот ÷ 1000 (выбытие ОС)
    165: {"source": "osv_kt", "accounts": ["0100", "0110"],
          "osv_div1000": True,
          "desc": "Выбыло ОС по первоначальной стоимости"},

    # 169 = ОСВ счёт 0100 Дебет оборот ÷ 1000 (поступление ОС)
    169: {"source": "osv_dt", "accounts": ["0100", "0110"],
          "osv_div1000": True,
          "desc": "Поступило ОС по первоначальной стоимости"},

    # ══ Глава 6: Инвестиции ═══════════════════════════════════════════
    # 180 = код 169 (инвестиции в ОС = поступление ОС)
    180: {"source": "calc", "formula": "169",
          "desc": "Инвестиции в основной капитал"},

    # 181 = ОСВ счёт 0400/04 Дебет оборот ÷ 1000 (непроизв. нефин. активы — земля)
    181: {"source": "osv_dt", "accounts": ["0400", "040", "04"],
          "osv_div1000": True,
          "desc": "Стоимость непроизведённых нефинансовых активов"},

    # 183 = код 169 − код 170 (приобретение б/у ОС у других лиц)
    183: {"source": "calc", "formula": "169-170",
          "desc": "Стоимость приобретённых ОС у других лиц"},

    # ══ Глава 9: Кадры ════════════════════════════════════════════════
    401: {"source": "personnel", "field": "avg_headcount_for_salary",
          "desc": "Численность для исчисления ЗП"},
    403: {"source": "personnel", "field": "total_wage_fund",
          "desc": "Начисленные доходы (ФОТ)"},
    404: {"source": "personnel", "field": "wage_fund_with_workbooks",
          "desc": "ФОТ работников с трудовыми книжками"},
    405: {"source": "personnel", "field": "headcount_with_workbooks_endyear",
          "desc": "Численность с трудкнижками на конец года"},
    409: {"source": "personnel", "field": "avg_headcount_with_workbooks",
          "desc": "Среднегодовая численность с трудкнижками"},
    411: {"source": "personnel", "field": "avg_external_parttime",
          "desc": "Внешние совместители (среднегодовые)"},
    412: {"source": "personnel", "field": "avg_gph_workers",
          "desc": "Работники по ГПХ (среднегодовые)"},
    413: {"source": "calc", "formula": "409+411+412",
          "desc": "Среднегодовая численность (вкл. совм. и ГПХ)"},
    416: {"source": "personnel", "field": "total_labor_costs",
          "desc": "Всего расходов на содержание рабочей силы"},

    # ══ Глава 10: Выплаты физлицам ════════════════════════════════════
    417: {"source": "personnel", "field": "interests_paid",   "desc": "Проценты"},
    418: {"source": "personnel", "field": "dividends_paid",   "desc": "Дивиденды"},
    419: {"source": "personnel", "field": "material_benefit", "desc": "Материальная выгода"},
    420: {"source": "personnel", "field": "material_aid",     "desc": "Материальная помощь"},
    421: {"source": "personnel", "field": "author_fees",      "desc": "Авторское вознаграждение"},
    422: {"source": "personnel", "field": "severance_pay",    "desc": "Выходное пособие"},
    423: {"source": "manual",    "desc": "Компенсационные выплаты"},
    424: {"source": "manual",    "desc": "Средства на обучение"},
}

# ── Контрольные соотношения ────────────────────────────────────────────
CONTROL_RATIOS = {
    "109 = 101+102+103": lambda r: abs(
        (r.get(109,{}).get("year") or 0) -
        ((r.get(101,{}).get("year") or 0) +
         (r.get(102,{}).get("year") or 0) +
         (r.get(103,{}).get("year") or 0))) < 1,
    "413 = 409+411+412": lambda r: abs(
        (r.get(413,{}).get("year") or 0) -
        ((r.get(409,{}).get("year") or 0) +
         (r.get(411,{}).get("year") or 0) +
         (r.get(412,{}).get("year") or 0))) < 1,
    "110 = 111+127": lambda r: abs(
        (r.get(110,{}).get("year") or 0) -
        ((r.get(111,{}).get("year") or 0) +
         (r.get(127,{}).get("year") or 0))) < 1,
}


class KorxonaProcessor:
    """Основной класс обработки данных для 1-korxona"""

    def __init__(self, f1_path=None, f2_path=None, osv_path=None, personnel_path=None):
        self.f1_path = Path(f1_path) if f1_path else None
        self.f2_path = Path(f2_path) if f2_path else None
        self.osv_path = Path(osv_path) if osv_path else None
        self.personnel_path = Path(personnel_path) if personnel_path else None

        self.f1_data = {}      # {row_code: {"begin": val, "end": val}}
        self.f2_data = {}      # {row_code: val}
        self.osv_data = {}     # {account: {"debit_turnover": val, "credit_turnover": val, ...}}
        self.ndfl_data = {}    # {row_code: value} — из НДФЛ-расчёта
        self.personnel_data = {}  # {field: val}
        self.results = {}      # {code: {"begin": val, "end": val, "year": val}}
        self.warnings = []

    # ──────────────────────────────────────────────────────────────────────
    # ПАРСЕРЫ
    # ──────────────────────────────────────────────────────────────────────

    def _find_header_row(self, df, keywords):
        """Найти строку-заголовок по ключевым словам"""
        for i, row in df.iterrows():
            row_str = " ".join(str(v) for v in row.values if pd.notna(v)).lower()
            if all(kw.lower() in row_str for kw in keywords):
                return i
        return None

    def _normalize_code(self, val):
        """Нормализовать код строки к строке '010', '040' и т.д."""
        if pd.isna(val):
            return None
        s = str(val).strip().split(".")[0].split(",")[0]
        s = re.sub(r"[^0-9]", "", s)
        return s.zfill(3) if s else None

    def _read_xltx(self, path):
        """Читает Excel/xltx — пробует лист list02 (формат 1С), иначе первый лист с данными."""
        path = str(path)
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        # list02 обычно содержит данные в формате 1С
        for sh in ['list02', 'list01'] + sheets:
            if sh not in sheets:
                continue
            try:
                df = pd.read_excel(path, sheet_name=sh, header=None, engine='openpyxl')
                # Есть ли числа?
                has_nums = any(isinstance(v,(int,float)) and pd.notna(v) and v not in (0,)
                               for row in df.values for v in row)
                if has_nums:
                    return df
            except Exception:
                continue
        return pd.read_excel(path, sheet_name=sheets[0], header=None, engine='openpyxl')

    def _find_code_col(self, df):
        """Находит индекс колонки с кодами строк (010, 020 и т.д.)"""
        best_col, best_count = 0, 0
        for ci in range(min(df.shape[1], 6)):
            count = sum(1 for v in df.iloc[:,ci]
                        if self._normalize_code(v) and len(self._normalize_code(v)) >= 2)
            if count > best_count:
                best_col, best_count = ci, count
        return best_col if best_count > 0 else None

    def parse_f1(self):
        """Парсинг Формы 1 (Бухгалтерский баланс).
        Поддерживает .xlsx и .xltx (шаблоны 1С с листами list01/list02).
        Структура: [Наименование, Код, Нач.года, Кон.года]
        """
        if not self.f1_path or not self.f1_path.exists():
            log.warning("Форма 1 не предоставлена, пропускаем.")
            return

        df = self._read_xltx(self.f1_path)
        log.info(f"Форма 1: загружено {len(df)} строк")

        code_col = self._find_code_col(df)
        if code_col is None:
            log.warning("Форма 1: не найдена колонка с кодами строк")
            return

        for _, row in df.iterrows():
            vals = list(row.values)
            if code_col >= len(vals):
                continue
            code = self._normalize_code(vals[code_col])
            if not code or len(code) > 4:
                continue
            # Числа правее колонки кода
            nums = [v for v in vals[code_col+1:code_col+5]
                    if isinstance(v,(int,float)) and pd.notna(v)]
            self.f1_data[code] = {
                "begin": nums[0] if len(nums) > 0 else 0,
                "end":   nums[1] if len(nums) > 1 else 0,
            }

        log.info(f"Форма 1: распознано {len(self.f1_data)} строк")

    def parse_f2(self):
        """Парсинг Формы 2 (ОПУ).
        Поддерживает .xlsx и .xltx.
        В xltx-формате 1С: код в col 2, доходы отч.года в col 5, расходы в col 6.
        В обычном xlsx: код рядом с первым числом.
        """
        if not self.f2_path or not self.f2_path.exists():
            log.warning("Форма 2 не предоставлена, пропускаем.")
            return

        df = self._read_xltx(self.f2_path)
        log.info(f"Форма 2: загружено {len(df)} строк")

        code_col = self._find_code_col(df)
        if code_col is None:
            log.warning("Форма 2: не найдена колонка с кодами строк")
            return

        for _, row in df.iterrows():
            vals = list(row.values)
            if code_col >= len(vals):
                continue
            code = self._normalize_code(vals[code_col])
            if not code or len(code) > 4:
                continue

            # Ф2 имеет структуру: доходы (столбец 5) и расходы (столбец 6)
            # Сохраняем оба значения отдельно
            nums = [(i, v) for i, v in enumerate(vals[code_col+1:code_col+8], code_col+1)
                    if isinstance(v,(int,float)) and pd.notna(v) and str(v) not in ('x',)]
            if len(nums) >= 2:
                # Первое ненулевое = доход, второе = расход
                income_val  = next((v for _,v in nums if v != 0), 0)
                # Ищем расход: обычно второй ненулевой
                nz = [v for _,v in nums if v != 0]
                expense_val = nz[1] if len(nz) >= 2 else 0
                self.f2_data[code] = {"income": income_val, "expense": expense_val}
            elif len(nums) == 1:
                self.f2_data[code] = {"income": nums[0][1], "expense": 0}

        log.info(f"Форма 2: распознано {len(self.f2_data)} строк")

    def parse_osv(self):
        """Парсинг ОСВ (Оборотно-сальдовая ведомость из 1С).

        Поддерживает форматы 1С:
          - .xls  (старый Excel, конвертируется через LibreOffice)
          - .xlsx (новый Excel)

        Структура строки ОСВ из 1С:
          col 0: "XXXX, Название счёта"
          col 2: Сальдо нач Дт
          col 3: Сальдо нач Кт
          col 4: Оборот Дт
          col 5: Оборот Кт
          col 7: Сальдо кон Дт
          col 8: Сальдо кон Кт

        Все значения в СУМАХ (делится на 1000 при извлечении, если osv_div1000=True).
        """
        if not self.osv_path or not self.osv_path.exists():
            log.warning("ОСВ не предоставлена, пропускаем.")
            return

        path = str(self.osv_path)

        # .xls — конвертируем через LibreOffice
        if path.lower().endswith(".xls"):
            import subprocess, tempfile, shutil
            tmp_dir = tempfile.mkdtemp()
            try:
                result = subprocess.run(
                    ["libreoffice", "--headless", "--convert-to", "xlsx",
                     "--outdir", tmp_dir, path],
                    capture_output=True, timeout=30
                )
                converted = [f for f in __import__("os").listdir(tmp_dir)
                             if f.endswith(".xlsx")]
                if converted:
                    path = __import__("os").path.join(tmp_dir, converted[0])
                    log.info(f"ОСВ: .xls конвертирован в {path}")
                else:
                    log.warning("ОСВ: конвертация .xls не удалась, пробуем напрямую")
            except Exception as e:
                log.warning(f"ОСВ: ошибка конвертации {e}")

        try:
            df = pd.read_excel(path, header=None, engine='openpyxl')
        except Exception:
            df = pd.read_excel(path, header=None)
        log.info(f"ОСВ: загружено {len(df)} строк")

        # Определяем формат: 1С (col 0 = "XXXX, Название") или обычный
        # Признак 1С: первая колонка содержит строки вида "0100, Основные средства"
        is_1c_format = any(
            re.match(r"^\d{2,4},", str(row.iloc[0]).strip())
            for _, row in df.iterrows()
            if pd.notna(row.iloc[0])
        )

        for _, row in df.iterrows():
            vals = list(row.values)
            acct_raw = str(vals[0]).strip() if pd.notna(vals[0]) else ""

            if not acct_raw or acct_raw in ("nan", "Итого", ""):
                continue

            # Код счёта: берём цифры до запятой (1С) или все начальные цифры
            if is_1c_format:
                acct = acct_raw.split(",")[0].strip()
                # Индексы колонок в формате 1С: нач_Дт=2, нач_Кт=3, об_Дт=4, об_Кт=5, кон_Дт=7, кон_Кт=8
                def g(idx):
                    if idx >= len(vals): return 0.0
                    v = vals[idx]
                    try: return float(v) if pd.notna(v) else 0.0
                    except: return 0.0
                self.osv_data[acct] = {
                    "begin_debit":    g(2), "begin_credit":  g(3),
                    "debit_turnover": g(4), "credit_turnover": g(5),
                    "end_debit":      g(7), "end_credit":    g(8),
                }
            else:
                # Обычный формат: счёт в col 0, затем числа подряд
                acct = re.sub(r"[^0-9]", "", acct_raw)
                if not acct or len(acct) < 2:
                    continue
                nums = []
                for v in vals[1:]:
                    try:
                        nums.append(float(v) if pd.notna(v) else 0.0)
                    except:
                        nums.append(0.0)
                if len(nums) >= 4:
                    self.osv_data[acct] = {
                        "begin_debit":    nums[0], "begin_credit":  nums[1],
                        "debit_turnover": nums[2], "credit_turnover": nums[3],
                        "end_debit":      nums[4] if len(nums) > 4 else 0,
                        "end_credit":     nums[5] if len(nums) > 5 else 0,
                    }

        log.info(f"ОСВ: распознано {len(self.osv_data)} счетов")

    def parse_personnel(self):
        """Парсинг данных по персоналу (отдельный Excel).
        Ожидаемый формат: два столбца — [Показатель / field_name, Значение]
        """
        if not self.personnel_path or not self.personnel_path.exists():
            log.warning("Файл персонала не предоставлен, пропускаем.")
            return

        df = pd.read_excel(self.personnel_path, header=None, engine='openpyxl')
        # Поддерживаем как словарный формат (ключ-значение), так и поиск по ключевым словам
        for _, row in df.iterrows():
            if len(row) >= 2:
                key = str(row.iloc[0]).strip().lower()
                val = row.iloc[1]
                if pd.notna(val):
                    self.personnel_data[key] = float(val) if isinstance(val, (int, float)) else val

        log.info(f"Персонал: загружено {len(self.personnel_data)} полей")

    # ──────────────────────────────────────────────────────────────────────
    # ВЫЧИСЛЕНИЕ ЗНАЧЕНИЙ
    # ──────────────────────────────────────────────────────────────────────

    def _get_f1(self, row_code, col="both"):
        data = self.f1_data.get(row_code.zfill(3), {})
        if col == "both":
            return {"begin": data.get("begin", 0), "end": data.get("end", 0)}
        return data.get(col, 0)

    def _get_f2(self, row_code, col="income"):
        """Получить значение из Ф2.
        col="income"  → первое числовое значение (доходы, столбец 5 в xltx)
        col="expense" → второе числовое значение (расходы, столбец 6 в xltx)
        """
        row_data = self.f2_data.get(row_code.zfill(3))
        if row_data is None:
            return 0
        if isinstance(row_data, dict):
            return row_data.get(col, 0) or 0
        return row_data or 0

    def _get_osv(self, accounts, col, div1000=False):
        """Получить сумму по счетам ОСВ.
        Ищет по точному совпадению и по первым символам (префикс).
        """
        total = 0.0
        for acct in accounts:
            for key, vals in self.osv_data.items():
                # Точное совпадение или префикс (напр. "9300" найдёт "9300", "9310" и т.д.)
                if key == acct or (len(acct) <= 4 and key.startswith(acct)):
                    total += vals.get(col, 0) or 0
        return total / 1000 if div1000 else total

    def _get_personnel(self, field):
        val = self.personnel_data.get(field)
        if val is None:
            for k, v in self.personnel_data.items():
                if field.lower() in k:
                    return float(v) if isinstance(v, (int, float)) else 0
        return float(val) if val is not None else 0

    def compute(self):
        """Вычислить все значения согласно маппингу (по шпаргалке TAVAT)"""
        computed = {}

        for code, cfg in MAPPING.items():
            src  = cfg["source"]
            desc = cfg["desc"]
            div  = cfg.get("osv_div1000", False)

            try:
                # ── Форма 1 ───────────────────────────────────────────────
                if src == "f1":
                    v = self._get_f1(cfg["row_code"], "both")
                    computed[code] = {"begin": v["begin"], "end": v["end"], "year": None}

                elif src == "f1_sum":
                    # Сумма нескольких строк Ф1 (нач/кон отдельно)
                    begin = sum(self._get_f1(rc,"both")["begin"] for rc in cfg["row_codes"])
                    end   = sum(self._get_f1(rc,"both")["end"]   for rc in cfg["row_codes"])
                    computed[code] = {"begin": begin, "end": end, "year": None}

                # ── Форма 2 ───────────────────────────────────────────────
                elif src == "f2":
                    col = cfg.get("col", "income")
                    computed[code] = {"year": self._get_f2(cfg["row_code"], col)}

                elif src == "f2_sum":
                    # Сумма строк Ф2, доходный столбец
                    total = sum(self._get_f2(rc, "income") for rc in cfg["row_codes"])
                    computed[code] = {"year": total}

                elif src == "f2_sum_exp":
                    # Сумма строк Ф2, расходный столбец (Код 111)
                    total = sum(self._get_f2(rc, "expense") for rc in cfg["row_codes"])
                    computed[code] = {"year": total}

                # ── ОСВ ───────────────────────────────────────────────────
                elif src == "osv_dt":
                    total = self._get_osv(cfg["accounts"], "debit_turnover",  div)
                    computed[code] = {"year": total}

                elif src == "osv_kt":
                    total = self._get_osv(cfg["accounts"], "credit_turnover", div)
                    computed[code] = {"year": total}

                elif src == "osv_end":
                    total = self._get_osv(cfg["accounts"], "end_debit", div)
                    computed[code] = {"year": total}

                # ── НДФЛ-расчёт ───────────────────────────────────────────
                elif src == "ndfl":
                    # Берётся из ndfl_data если есть, иначе 0
                    val = self.ndfl_data.get(cfg.get("ndfl_row", ""), 0) or 0
                    if cfg.get("ndfl_div1000"):
                        val = val / 1000
                    computed[code] = {"year": val}

                # ── Персонал ──────────────────────────────────────────────
                elif src == "personnel":
                    computed[code] = {"year": self._get_personnel(cfg["field"])}

                # ── Ручной ввод ───────────────────────────────────────────
                elif src == "manual":
                    computed[code] = {"year": 0}
                    self.warnings.append(f"Код {code} ({desc}): требует ручного ввода")

                # ── Формула ───────────────────────────────────────────────
                elif src == "calc":
                    computed[code] = {"year": None, "_formula": cfg["formula"]}

            except Exception as e:
                log.warning(f"Код {code} ({desc}): ошибка — {e}")
                computed[code] = {"year": 0}

        # Второй проход — вычисляем формулы (поддержка + и -)
        for code, vals in computed.items():
            if vals.get("_formula"):
                formula = vals["_formula"]
                # Парсим формулу вида "101+102+103" или "169-170"
                total = 0
                tokens = re.findall(r"([+-]?)(\d+)", formula)
                for sign, num_str in tokens:
                    dep_code = int(num_str)
                    dep_val  = (computed.get(dep_code, {}).get("year") or 0)
                    if sign == "-":
                        total -= dep_val
                    else:
                        total += dep_val
                computed[code] = {"year": total}

        self.results = computed
        return computed

    # ──────────────────────────────────────────────────────────────────────
    # ВАЛИДАЦИЯ
    # ──────────────────────────────────────────────────────────────────────

    def validate(self):
        """Проверка контрольных соотношений"""
        errors = []
        for name, check_fn in CONTROL_RATIOS.items():
            try:
                if not check_fn(self.results):
                    val_parts = {}
                    for p in re.findall(r"\d+", name):
                        c = int(p)
                        val_parts[c] = (self.results.get(c, {}).get("year") or 0)
                    errors.append({"ratio": name, "values": val_parts})
            except Exception as e:
                log.warning(f"Ошибка контрольного соотношения {name}: {e}")
        return errors

    # ──────────────────────────────────────────────────────────────────────
    # ЭКСПОРТ В ШАБЛОН
    # ──────────────────────────────────────────────────────────────────────

    def fill_template(self, template_path: str, output_path: str):
        """Заполнить Excel-шаблон Госкомстата вычисленными значениями.

        Алгоритм поиска ячейки:
        1. Сканируем все строки листа
        2. Ищем ячейку со значением == код строки (целое число или строка)
        3. Значение вставляем в соседние ячейки (по колонкам)
        """
        wb = load_workbook(template_path)

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    # Попытка трактовать значение ячейки как код строки
                    try:
                        cell_code = int(str(cell.value).strip().split(".")[0])
                    except (ValueError, AttributeError):
                        continue

                    if cell_code not in self.results:
                        continue

                    mapping_cfg = MAPPING[cell_code]
                    result = self.results[cell_code]
                    col_idx = cell.column

                    # Определяем колонки для заполнения
                    if mapping_cfg.get("col") == "both" or ("begin" in result and "end" in result):
                        # Ищем 2 числовых ячейки правее: начало и конец года
                        filled = 0
                        for offset in range(1, 10):
                            neighbor = sheet.cell(row=cell.row, column=col_idx + offset)
                            if neighbor.value is None or isinstance(neighbor.value, (int, float)):
                                if filled == 0 and result.get("begin") is not None:
                                    neighbor.value = result["begin"]
                                    filled += 1
                                elif filled == 1 and result.get("end") is not None:
                                    neighbor.value = result["end"]
                                    filled += 1
                                    break
                    else:
                        # Ищем первую пустую числовую ячейку правее
                        for offset in range(1, 10):
                            neighbor = sheet.cell(row=cell.row, column=col_idx + offset)
                            if neighbor.value is None or isinstance(neighbor.value, (int, float)):
                                neighbor.value = result.get("year") or 0
                                break

        wb.save(output_path)
        log.info(f"Шаблон сохранён: {output_path}")

    # ──────────────────────────────────────────────────────────────────────
    # ПОЛНЫЙ ЦИКЛ
    # ──────────────────────────────────────────────────────────────────────

    def run(self, template_path=None, output_path=None):
        """Запустить полный цикл обработки"""
        self.parse_f1()
        self.parse_f2()
        self.parse_osv()
        self.parse_personnel()
        self.compute()
        errors = self.validate()

        if template_path and output_path:
            self.fill_template(template_path, output_path)

        return {
            "results": self.results,
            "validation_errors": errors,
            "warnings": self.warnings,
        }
